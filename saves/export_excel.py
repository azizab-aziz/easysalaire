import csv
import sys
import subprocess

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─── Read CSV ─────────────────────────────
rows = []
with open("C:\\EasySalaire\\saves\\employes.csv", "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    for row in reader:
        rows.append(row)

# ─── Create workbook ──────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employes"

# ─── Colors ───────────────────────────────
NAVY   = "1A2B4A"
WHITE  = "FFFFFF"
LIGHT  = "EAF2FF"
GREEN  = "16A34A"
RED    = "DC2626"

# ─── Border ───────────────────────────────
thin   = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── Header row ───────────────────────────
headers = [
    "ID","Nom", "Prenom", "Poste",
    "Salaire Base", "Heures Sup", "Prime",
    "CNSS", "IR", "Salaire Net", "Periode","Bulletin"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font      = Font(bold=True, color=WHITE, size=11)
    cell.fill      = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border

ws.row_dimensions[1].height = 30

# ─── Data rows ────────────────────────────
for row_idx, row in enumerate(rows[1:], 2):
    bg = LIGHT if row_idx % 2 == 0 else WHITE
    fill = PatternFill("solid", fgColor=bg)

    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx)

        if col_idx == 11:
            # Bulletin column → keep as text N°001
            cell.value = value.strip()
            cell.font  = Font(color="1E293B", bold=True, size=10)
        elif col_idx == 9:
            try:
                cell.value = float(value)
                cell.font  = Font(color=GREEN, bold=True, size=10)
            except:
                cell.value = value
                cell.font  = Font(color="1E293B", size=10)
        elif col_idx in [7, 8]:
            try:
                cell.value = float(value)
                cell.font  = Font(color=RED, size=10)
            except:
                cell.value = value
                cell.font  = Font(color="1E293B", size=10)
        elif col_idx == 10:
            # Periode column → text
            cell.value = value.strip()
            cell.font  = Font(color="1E293B", size=10)
        else:
            try:
                cell.value = float(value)
                cell.font  = Font(color="1E293B", size=10)
            except:
                cell.value = value
                cell.font  = Font(color="1E293B", size=10)

        cell.fill      = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = border

    ws.row_dimensions[row_idx].height = 22

# ─── Auto-fit column widths ───────────────
for col in ws.columns:
    max_len   = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 6

# ─── Freeze header + auto filter ──────────
ws.freeze_panes    = "A2"
ws.auto_filter.ref = ws.dimensions

# ─── Save ─────────────────────────────────
wb.save("C:\\EasySalaire\\saves\\employes.xlsx")
print("Excel exported successfully")